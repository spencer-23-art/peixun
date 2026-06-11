import openpyxl
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('06.03.xlsx')
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [str(cell.value) if cell.value is not None else '' for cell in row]
        print(' | '.join(vals))
    print()
